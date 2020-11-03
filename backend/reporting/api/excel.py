import re
from typing import List, Dict

import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment
from openpyxl.styles.colors import WHITE
from openpyxl.utils import get_column_letter as to_letter
from django.template.defaultfilters import pluralize

from .models import Status, Validation
from utils import intel_calendar

BOLD_FONT = Font(bold=True)

# Default table style (zebra + color + header sort/filters)
MEDIUM_STYLE = TableStyleInfo(name='TableStyleMedium6', showRowStripes=True)


def do_report(data: pd.DataFrame=None, extra=None, report_name=''):
    validations = extra or []
    ct = data

    wb = Workbook()
    ws = wb.active
    year, work_week, weekday = intel_calendar.ww_date()
    rdate = f'{report_name}: {year} ww{work_week}.{weekday}'
    ws['A2'] = rdate
    ws['A2'].font = Font(bold=True, size=14)
    ws.merge_cells('A2:G2')

    ws['A3'] = 'Report generated based on validations:'
    ws.merge_cells('A3:G3')

    # List of selected validations
    c_row = 4
    formatted_validations = []
    for v in validations:
        formatted_validations.append(f'{v.name} ({v.platform.name}, {v.env.name}, {v.os.name})')

    for v in formatted_validations:
        c = ws.cell(row=c_row, column=1)
        c.value = v
        c.font = BOLD_FONT
        ws.merge_cells(start_row=c_row, start_column=1, end_row=c_row, end_column=7)
        c_row += 1
    c_row += 1

    # Starting table creation
    table_row_start, table_columns = c_row + 1, 1
    table_row_end = c_row + 1

    # DataFrame to cells
    rows = dataframe_to_rows(ct, index=True, header=True)
    for r_id, row in enumerate(rows, 1):
        table_columns = len(row)
        c_row += 1
        table_row_end = c_row
        for c_id, value in enumerate(row, 1):
            ws.cell(row=c_row, column=c_id, value=value)

    # Insert first column header name
    ws.cell(row=table_row_start, column=1, value='Group name')

    # First column
    max_groups_len = max(map(len, list(ct.index)))

    # Format first and last row in "table"
    column_names = list(ct.columns)  # not run, error, failed passed total, pass-rate
    for c_id in range(1, table_columns + 1):
        ws.cell(row=table_row_start, column=c_id).font = Font(color=WHITE)
        ws.cell(row=table_row_end, column=c_id).font = BOLD_FONT

        # Set columns default width by text len
        if c_id != 1:
            ws.column_dimensions[to_letter(c_id)].width = len(column_names[c_id - 2]) + 4   # padding for filter button
        else:
            ws.column_dimensions[to_letter(c_id)].width = max_groups_len

    table = Table(
        ref=f'{to_letter(1)}{table_row_start}:{to_letter(table_columns)}{table_row_end - 1}',
        displayName='Results', tableStyleInfo=MEDIUM_STYLE)
    ws.add_table(table)

    # Remove empty row where DataFrame index name should be located
    ws.delete_rows(table_row_start + 1, 1)

    return wb


def do_comparison_report(ct: pd.DataFrame) -> Workbook:
    statuses = Status.objects.all().values_list('test_status', flat=True)

    wb = Workbook()
    ws = wb.active

    year, work_week, weekday = intel_calendar.ww_date()
    rdate = f'Validations comparison report: {year} ww{work_week}.{weekday}'
    ws['A2'] = rdate
    ws['A2'].font = Font(bold=True, size=14)
    ws.merge_cells('A2:G2')

    c_row = 3

    # Starting table creation
    table_row_start, table_columns = c_row + 1, 1
    table_row_end = c_row + 1

    # DataFrame to cells
    rows = dataframe_to_rows(ct, index=True, header=True)
    for r_id, row in enumerate(rows, 1):
        table_columns = len(row)
        c_row += 1
        table_row_end = c_row
        for c_id, value in enumerate(row, 1):
            ws.cell(row=c_row, column=c_id, value=value)
            if r_id == table_row_start:
                ws.cell(row=c_row, column=c_id).alignment = Alignment(wrap_text=True)

            font = Font()
            # coloring statuses
            if value in statuses:
                if value == 'Passed':
                    font = Font(color='2E7D32')
                elif value == 'Failed':
                    font = Font(color='B71C1C')
                elif value == 'Error':
                    font = Font(color='D84315')
                elif value == 'Blocked':
                    font = Font(color='616161')
                elif value == 'Skipped':
                    font = Font(color='00838F')
                elif value == 'Canceled':
                    font = Font(color='3E2723')
                font.bold = True
                ws.cell(row=c_row, column=c_id).alignment = Alignment(horizontal='center')
            ws.cell(row=c_row, column=c_id).font = font

    # Insert first column header name
    ws.cell(row=table_row_start, column=1, value='Item name')

    # First column
    max_first_column_len = max(map(len, list(ct.index)))

    # Format first and last row in "table"
    for c_id in range(1, table_columns + 1):
        ws.cell(row=table_row_start, column=c_id).font = Font(color=WHITE)
        ws.column_dimensions[to_letter(c_id)].width = max_first_column_len

    table = Table(
        ref=f'{to_letter(1)}{table_row_start}:{to_letter(table_columns)}{table_row_end - 1}',
        displayName='Results', tableStyleInfo=MEDIUM_STYLE)
    ws.add_table(table)

    # Remove empty row where DataFrame index name should be located
    ws.delete_rows(table_row_start + 1, 1)

    return wb


def do_indicator_report(data, validation, mappings, mode):
    statuses = ['total', 'passed', 'failed', 'blocked', 'executed', 'not_run']
    wb = Workbook()

    def report_sheet(data, validation, mappings, mode):
        # create or select current sheet, rename it
        if mode == 'combined':
            ws = wb.active
            ws.title = 'Combined'
        else:
            ws = wb.create_sheet(compose_sheet_name(mappings[0].codec))

        # Caption
        year, work_week, weekday = intel_calendar.ww_date()
        rdate = f'Indicator report: {year} ww{work_week}.{weekday}'
        ws['A2'] = rdate
        ws['A2'].font = Font(bold=True, size=12)
        ws.merge_cells('A2:E2')
        ws['A3'] = f'Report generated for validation: {validation.name}'
        ws['A3'].font = Font(bold=True, size=12)
        ws.merge_cells('A3:E3')
        ws['A4'] = 'Feature Mapping Tables used:'
        ws['A5'], ws['B5'] = 'Name', 'Author'
        c_row = 6

        for mapping in mappings:
            ws.cell(row=c_row, column=1, value=mapping.name)
            ws.cell(row=c_row, column=2, value=mapping.owner.username)
            c_row += 1
        c_row += 1

        # Calculate table dimensions
        features_tuples = []
        for milestone_data in data['items'].values():
            features_tuples.extend(milestone_data.items())

        table_row_start = c_row
        table_row_end = len(features_tuples) + len(data['items'].keys())
        col_width = {}

        # Header row
        headers = ['Feature', 'Total', 'Passed', 'Failed', 'Blocked', 'Executed', 'Not Run']
        for c_id, header in enumerate(headers, 1):
            # set width for columns except features
            if c_id > 1:
                col_width[c_id] = len(header)

            ws.cell(row=c_row, column=c_id, value=header).font = BOLD_FONT
        c_row += 1

        # Items rows
        for milestone, features_data in data['items'].items():
            ws.cell(row=c_row, column=1, value=milestone)

            c_row += 1
            for feature, f_dict in features_data.items():
                ws.cell(row=c_row, column=1, value=feature)

                col_width.setdefault(1, 0)
                if len(feature) > col_width[1]:
                    col_width[1] = len(feature)

                for c_id, key in enumerate(statuses, 2):
                    ws.cell(row=c_row, column=c_id, value=f_dict[key])
                c_row += 1

        # Total row
        ws.cell(row=c_row, column=1, value='Total').font = BOLD_FONT
        for c_id, key in enumerate(statuses, 2):
            ws.cell(row=c_row, column=c_id, value=data['total'][key])

        # set column width
        for col_ind in col_width:
            ws.column_dimensions[to_letter(col_ind)].width = col_width[col_ind] + 4     # + padding for filter button

        # Apply table style (zebra + color + header sort/filters)
        if mode == 'combined':
            display_name = 'Combined'
        else:
            display_name = f'Single_{mappings[0].id}'
        table = Table(
            ref=f'A{table_row_start}:G{table_row_start + table_row_end + 1}',
            displayName=display_name, tableStyleInfo=MEDIUM_STYLE)
        ws.add_table(table)

    if mode == 'combined':
        report_sheet(data, validation, mappings, 'combined')
    else:
        for mapping in mappings:
            report_sheet(data[mapping.id], validation, [mapping], 'single')
        # delete default empty sheet
        del wb[wb.sheetnames[0]]
    return wb


def compose_sheet_name(codec):
    """
    Return mapping codec name, formatted for Excel sheet name limitations:
    Banned symbols: "\", "/", "*", "[", "]", ":", "?"
    Length limit: 31
    """
    name = re.sub(r'[\\/*\[\]:?]', '_', codec.name)
    if len(name) > 31:
        name = f'{name[:27]}#{codec.id}'
    return name


def do_issues_report(pk, failed_groups, error_groups):
    wb = Workbook()
    ws = wb.active

    val = Validation.objects.get(pk=pk)
    ws['B1'] = f"Issues report for '{val.name}'"
    ws['B1'].font = Font(bold=True, size=12)

    total_failed = 0
    for failed_list in failed_groups.values():
        total_failed += len(failed_list)

    row = 3
    if total_failed > 0:
        ws[f'B{row}'] = f'{total_failed} Test Item{pluralize(total_failed)} with status Failed'
        ws[f'B{row}'].font = Font(bold=True, size=12)

        row += 2
        for error_feature, failed_list in failed_groups.items():
            ws[f'A{row}'] = error_feature
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1

            for item in failed_list:
                ws[f'A{row}'] = item['ti']
                ws[f'B{row}'] = item['err']
                row += 1

    total_error = 0
    for error_list in error_groups.values():
        total_error += len(error_list)

    if total_error > 0:
        if total_failed != 0:
            row += 2

        ws[f'B{row}'] = f'{total_error} Test Item{pluralize(total_error)} with status Error'
        ws[f'B{row}'].font = Font(bold=True, size=12)

        row += 2
        for error_feature, error_list in error_groups.items():
            ws[f'A{row}'] = error_feature
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1

            for item in error_list:
                ws[f'A{row}'] = item['ti']
                ws[f'B{row}'] = item['err']
                row += 1

    if total_failed == 0 and total_error == 0:
        ws['A3'] = 'There are no Test Items with Failed or Error statuses'
        ws['A3'].font = Font(bold=True, size=12)
    else:
        ws.column_dimensions['A'].width = 24

    return wb