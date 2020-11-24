import logging
import requests

from io import BytesIO
from typing import Any, Union, Dict
from requests.auth import HTTPBasicAuth

from django.conf import settings
from django.core.files.uploadedfile import InMemoryUploadedFile
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

log = logging.getLogger(__name__)

NAME_MAPPING = {
    'buildversion': 'driverName',
    'item name': 'itemName',
    'args': 'itemArgs',
    'component': 'componentName',
    'feature': 'feature',
    'execution start time': 'execStart',
    'execution end time': 'execEnd',
    'environment': 'envName',
    'operating system': 'osVersion',
    'operating system family': 'osName',
    'platform': 'platformName',
    'result key': 'resultKey',
    'status': 'status',
    'test run': 'testRun',
    'test session': 'testSession',
    'url': 'resultURL',
    'reason': 'reason',
    'vertical': 'vertical',
    'mapped component': 'mappedComponent',
}

REVERSE_NAME_MAPPING = {value: key for key, value in NAME_MAPPING.items()}


def open_excel_file(path: Union[str, InMemoryUploadedFile], outcome: Any = None) -> Workbook:
    try:
        if (isinstance(path, str)
                and path.startswith('http')
                and 'artifactory/gta-results/excel' in path):
            return open_remote_excel_file(path)
        return load_workbook(path)
    except Exception as e:
        message = f"Failed to open workbook: {getattr(e, 'message', repr(e))}"

        log.warning(message)
        if outcome:
            outcome.add_workbook_error(message)


def open_remote_excel_file(url_path: str) -> Workbook:
    """
    Open remote excel file as workbook by url from buffer data
    :param url_path: artifactory location of the exel report
    :return: encoded excel workbook
    """
    raw_data = requests.get(url=url_path,
                            auth=HTTPBasicAuth(settings.ARTIFACTORY_API_USER,
                                               settings.ARTIFACTORY_API_PASSWORD))
    return load_workbook(BytesIO(raw_data.content))


# generator that stops row iteration on first empty row
def non_empty_row(rows):
    while True:
        try:
            row = next(rows)
        except StopIteration:
            break

        if set(cell.value for cell in row) == {None}:
            break

        yield row


class SheetMapping:
    sheet: Worksheet = None
    column_mapping: Dict[str, int] = {}

    def set_from_workbook(self, workbook, outcome):
        for title in [workbook.worksheets[0].title, 'best', 'all']:
            if title in workbook.sheetnames:
                sheet = workbook[title]
                self._create_column_mapping(workbook[title])

                if len(self.column_mapping) == len(NAME_MAPPING):
                    self.sheet = sheet
                    return True

        self._notify_missing_columns(outcome)
        return False

    def get_column_values(self, column_name=None, index=None):
        if column_name == index == None:
            return []

        rows = self.sheet.rows
        next(rows)

        if index is None:
            index = self.column_mapping[column_name]

        return [row[index].value for row in non_empty_row(rows)]

    def _notify_missing_columns(self, outcome):
        existing_columns = set(self.column_mapping.keys())
        missing_columns = ', '.join(
                key for key, value in NAME_MAPPING.items() if value not in existing_columns
        )

        outcome.add_missing_columns_error(missing_columns)

    def _create_column_mapping(self, sheet):
        captions = next(sheet.rows)
        column_mapping = dict()

        for cell in captions:
            if cell.value is None:
                continue

            mapped_name = NAME_MAPPING.get(str(cell.value).lower(), None)
            if mapped_name is not None:
                column_mapping[mapped_name] = cell.column - 1

        self.column_mapping = column_mapping
