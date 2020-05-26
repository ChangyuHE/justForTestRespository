import json
import logging

from datetime import datetime
from openpyxl import Workbook
from pathlib import Path
from django.core.serializers.json import DjangoJSONEncoder

log = logging.getLogger(__name__)

COLUMNS = [
    'buildVersion',
    'item name',
    'args',
    'component',
    'execution start time',
    'execution end time',
    'environment',
    'operating system',
    'operating system family',
    'platform',
    'result key',
    'status',
    'test run',
    'test session',
    'url',
    'reason',
]

def create_file(json_data_filename):
    filepath = Path(*__package__.split('.'), json_data_filename)
    workbook = Workbook()
    sheet = workbook.active

    # workbook table content is stored in external json file
    with open(filepath, 'r') as fp:
        json_data = fp.read()

    # create table captions
    for i in range(len(COLUMNS)):
        cell = sheet.cell(row=1, column=i+1)
        cell.value = COLUMNS[i]

    existing_columns = [False for _ in range(len(COLUMNS))]

    # fill in table
    row = 2
    for data in json.loads(json_data):
        for key, value in data.items():
            if key in COLUMNS:
                index = COLUMNS.index(key)
                column = index + 1
                existing_columns[index] = True

                if key in ['execution start time', 'execution end time']:
                    try:
                        value = datetime.fromisoformat(value)
                    except (ValueError, TypeError):
                        pass

                sheet.cell(row=row, column=column).value = value
        row += 1

    # delete empty columns
    for i in reversed(range(len(COLUMNS))):
        if not existing_columns[i]:
            sheet.delete_cols(i + 1)

    return workbook

def create_empty_workbook():
    workbook = Workbook()
    sheet = workbook.active

    # create table captions
    for i in range(len(COLUMNS)):
        cell = sheet.cell(row=1, column=i+1)
        cell.value = COLUMNS[i]

    sheet.cell(3, 1).value = None

    return workbook

def export_to_json(sheet, max_row):
    all_data = list()
    rows = sheet.rows
    next(rows)

    try:
        for _ in range(max_row):
            data = dict()
            for cell in next(rows):
                name = sheet.cell(row=1, column=cell.column).value
                if name in COLUMNS:
                    data[name] = cell.value
            all_data.append(data)

    except StopIteration:
        pass

    json_data =  json.dumps(all_data, sort_keys=True, indent=4, cls=DjangoJSONEncoder)
    return json_data
