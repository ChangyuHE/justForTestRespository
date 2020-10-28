import logging
from itertools import combinations
from collections import defaultdict

from typing import Union, Optional, List, Dict, Tuple

from django.db import transaction
from django.db.utils import DatabaseError, IntegrityError
from django.core.exceptions import ObjectDoesNotExist, MultipleObjectsReturned
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, render

from rest_framework import generics, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import FileUploadParser
from rest_framework.exceptions import ParseError

from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter as to_letter
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from api.models import FeatureMapping, FeatureMappingRule, Milestone, Feature, TestScenario
from api.forms import FeatureMappingFileForm
from api.serializers import FeatureMappingSimpleSerializer, FeatureMappingSerializer, FeatureMappingRuleSerializer, \
    FeatureMappingSimpleRuleSerializer, MilestoneSerializer, FeatureSerializer, TestScenarioSerializer

from utils.api_logging import LoggingMixin, get_user_object
from utils.api_helpers import get_datatable_json, DefaultNameOrdering
from utils import api_helpers

log = logging.getLogger(__name__)


class ValidationException(Exception):
    pass


class FeatureMappingPostView(LoggingMixin, APIView):
    """ Excel file import view """
    parser_class = (FileUploadParser,)

    def post(self, request):
        if 'file' not in request.data:
            raise ParseError("'file' parameter is missing in form data.")

        fm_serializer = FeatureMappingSimpleSerializer(data=request.data)
        if not fm_serializer.is_valid():
            for field, errors in fm_serializer.errors.items():
                if field == 'non_field_errors':
                    # we have only one non field checker so we may have only
                    # one non field error
                    fm_serializer.errors['non_field_errors'].clear()
                    fm_serializer.errors['non_field_errors'].append(
                        f"You already have FMT named '{request.data['name']}'"
                    )
                    break
            return Response(
                {'errors': fm_serializer.errors},
                status=status.HTTP_422_UNPROCESSABLE_ENTITY
            )

        errors = import_feature_mapping(request.data['file'], fm_serializer)
        if errors:
            return Response({'errors': errors}, status=status.HTTP_422_UNPROCESSABLE_ENTITY)
        return Response(status=status.HTTP_201_CREATED)


def check_mappings(ids):
    # prevent selection of mappings with the same codec and platform
    codecs = FeatureMapping.objects.filter(id__in=ids).values('codec', 'platform').distinct()
    if len(ids) != codecs.count():
        return False
    return True


class FeatureMappingConflictCheckView(LoggingMixin, APIView):
    """ Check for conflicts through selected mappings """
    def get(self, request):
        ids = request.query_params.get('ids', '').split(',')
        return Response(check_mappings(ids))


def check_rules_conflicts(ids_list, scenario_name, issues):
    # ids_list example: [[1,2,3], None, [4,5]]
    # first check - None with other ids in rules for one scenario
    for ids_tuple in combinations(ids_list, 2):
        if None in ids_tuple and any(e is not None for e in ids_tuple):
            issues[scenario_name].append('No empty Ids allowed if there are other rules with specified Ids exist')
            break

    # second check - same ids in rules for one scenario
    ids_wo_none = [ids.split(',') for ids in ids_list if ids is not None]
    for ids_tuple in combinations(ids_wo_none, 2):
        if set(ids_tuple[0]) & set(ids_tuple[1]):
            issues[scenario_name].append('There must be no same ids in different rules for the same scenario')
            break
    return issues


class FeatureMappingRulesConflictCheckView(LoggingMixin, APIView):
    """ Check for conflicts through mapping rules """
    def get(self, request, pk, *args, **kwargs):
        issues = defaultdict(list)
        rule_to_exclude = None

        # prepare data for edited/created rule
        new_ids = request.query_params.get('new_ids')
        if new_ids == 'none':
            new_ids = None
        scenario_id = request.query_params.get('scenario_id')

        # rule.id for edit or 'none' for new rule creation
        rule_id = request.query_params.get('rule_id')
        if rule_id != 'none':
            rule_to_exclude = rule_id

        scenario_name = TestScenario.objects.get(id=scenario_id).name
        rules = FeatureMappingRule.objects.filter(scenario_id=scenario_id, mapping_id=pk)
        # exclude ids for edited rule, because we passed new ones
        if rule_to_exclude:
            rules = rules.exclude(id=rule_to_exclude)

        # add new items to examine list
        ids_list = list(rules.values_list('ids', flat=True))
        ids_list.append(new_ids)

        issues = check_rules_conflicts(ids_list, scenario_name, issues)
        return Response(issues)


class FeatureMappingListView(LoggingMixin, DefaultNameOrdering, generics.ListAPIView):
    """ List of available FeatureMappings filtered by owner/platform/os/component"""
    queryset = FeatureMapping.objects.all()
    serializer_class = FeatureMappingSerializer
    filterset_fields = ['name', 'owner', 'platform', 'os', 'component', 'public', 'official']


class FeatureMappingCloneView(LoggingMixin, APIView):
    """
    post: Clone and edit selected feature mapping table
    """
    def post(self, request, pk):
        fmt = get_object_or_404(FeatureMapping, pk=pk)
        try:
            with transaction.atomic():
                mapping_rules = fmt.rules.all()
                cloned_fmt = fmt
                cloned_fmt.pk = None
                cloned_fmt.official = False
                cloned_fmt.owner = get_user_object(request)
                cloned_fmt.name = request.data['name']
                cloned_fmt.save()

                # obtain backward related rules to clone
                for rule in mapping_rules:
                    rule.pk = None
                    rule.mapping_id = cloned_fmt.id
                    rule.save()
                    cloned_fmt.rules.add(rule)
        except DatabaseError as e:
            return Response({'errors': e}, status=status.HTTP_400_BAD_REQUEST)

        # trick to allow public foreign tables cloning without changing their names to pass in serializer
        data = {k: v for k, v in request.data.items() if k != 'owner'}
        fm_serializer = FeatureMappingSimpleSerializer(cloned_fmt, data=data, partial=True)
        if not fm_serializer.is_valid():
            return Response({'errors': fm_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

        response_data = FeatureMappingSerializer(fm_serializer.save()).data
        return Response(response_data, status=status.HTTP_201_CREATED)


class FeatureMappingTableView(LoggingMixin, DefaultNameOrdering, generics.ListAPIView):
    """ FeatureMapping table view formatted for DataTable """
    queryset = FeatureMapping.objects.all()
    serializer_class = FeatureMappingSerializer
    filterset_fields = ['name', 'owner', 'platform', 'os', 'component', 'public', 'official']

    def get(self, request, *args, **kwargs):
        public = request.GET.get('public')
        exclude = []
        if not public:
            exclude = ['owner']
        return get_datatable_json(self, exclude=exclude)


class FeatureMappingDetailsView(LoggingMixin, generics.RetrieveDestroyAPIView, api_helpers.UpdateWOutputAPIView):
    """ FeatureMapping single object management """
    queryset = FeatureMapping.objects.all()
    serializer_class = FeatureMappingSimpleSerializer
    serializer_output_class = FeatureMappingSerializer


class FeatureMappingExportView(LoggingMixin, APIView):
    """ Export mapping as excel file """
    def get(self, request, *args, **kwargs):
        pk = kwargs.get('pk')
        if not pk:
            return Response({'details': 'no pk provided'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            mapping = FeatureMapping.objects.get(pk=pk)
        except ObjectDoesNotExist:
            return Response({'details': f'Could not find mapping by id {pk}'}, status=status.HTTP_404_NOT_FOUND)
        except MultipleObjectsReturned:
            return Response({'details': f'More than one objects found by id {pk}'}, status=status.HTTP_400_BAD_REQUEST)

        workbook = export_mapping(mapping)
        filename = f'FMT_{mapping.name}_{datetime.now():%Y-%m-%d_%H:%M:%S}.xlsx'
        response = HttpResponse(save_virtual_workbook(workbook), content_type="application/ms-excel")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


def export_mapping(mapping):
    wb = Workbook()
    ws = wb.active

    # first row (headers)
    for col, value in enumerate(('milestone', 'feature', 'scenario', 'ids'), start=1):
        ws.cell(row=1, column=col).value = value

    current_row = 2
    col_width = dict()
    # fill rows with data
    for values in FeatureMappingRule.objects.filter(mapping=mapping) \
            .values_list('milestone__name', 'feature__name', 'scenario__name', 'ids'):
        for col, value in enumerate(values, start=1):
            # collect max width per column
            col_width.setdefault(col, 0)
            if value and len(value) > col_width[col]:
                col_width[col] = len(value)

            ws.cell(row=current_row, column=col).value = value
        current_row += 1

    # set default column width
    for col_ind in col_width:
        ws.column_dimensions[to_letter(col_ind)].width = col_width[col_ind] + 3

    medium_style = TableStyleInfo(name='TableStyleMedium6', showRowStripes=True)

    table = Table(
        ref=f'A1:D{current_row - 1}', displayName='FMT', tableStyleInfo=medium_style)
    ws.add_table(table)

    return wb


# FeatureMapping Rules views

class FeatureMappingRuleDetailsView(LoggingMixin, generics.RetrieveDestroyAPIView, api_helpers.UpdateWOutputAPIView):
    """ FeatureMappingRule single object management """
    queryset = FeatureMappingRule.objects.all()
    serializer_class = FeatureMappingSimpleRuleSerializer
    serializer_output_class = FeatureMappingRuleSerializer


class FeatureMappingRuleView(LoggingMixin, generics.ListAPIView, api_helpers.CreateWOutputApiView):
    """
        get: List FeatureMappingRule objects according to filters
        post: Create FeatureMappingRule object using simple serializer, output with fill data
    """
    queryset = FeatureMappingRule.objects.all()
    serializer_output_class = FeatureMappingRuleSerializer
    filterset_fields = ['milestone', 'feature', 'scenario', 'mapping', 'total']

    def get_serializer_class(self):
        if self.request.method == 'GET':
            return FeatureMappingRuleSerializer
        return FeatureMappingSimpleRuleSerializer


class FeatureMappingRuleTableView(LoggingMixin, generics.ListAPIView):
    """ FeatureMappingRule table view formatted for DataTable """
    queryset = FeatureMappingRule.objects.all()
    serializer_class = FeatureMappingRuleSerializer
    filterset_fields = ['milestone', 'feature', 'scenario', 'mapping', 'total']

    def get(self, request, *args, **kwargs):
        return get_datatable_json(self)


# Milestone views
class FeatureMappingMilestoneView(LoggingMixin, DefaultNameOrdering, generics.ListAPIView):
    """ List FeatureMapping Milestone objects """
    queryset = Milestone.objects.all()
    serializer_class = MilestoneSerializer
    filterset_fields = ['name']


class FeatureMappingMilestoneDetailsView(LoggingMixin, generics.RetrieveUpdateDestroyAPIView):
    """ FeatureMapping Milestone single object management """
    queryset = Milestone.objects.all()
    serializer_class = MilestoneSerializer


class FeatureMappingMilestoneTableView(LoggingMixin, DefaultNameOrdering, generics.ListAPIView):
    """ FeatureMapping Milestone table view formatted for DataTable """
    queryset = Milestone.objects.all()
    serializer_class = MilestoneSerializer
    filterset_fields = ['name']

    def get(self, request, *args, **kwargs):
        return get_datatable_json(self, actions=False)


# Feature views
class FeatureMappingFeatureView(LoggingMixin, DefaultNameOrdering, generics.ListCreateAPIView):
    """
        get: List FeatureMapping Feature objects according to filters
        post: Create FeatureMapping Feature object
    """
    queryset = Feature.objects.all()
    serializer_class = FeatureSerializer
    filterset_fields = ['name']


class FeatureMappingFeatureDetailsView(LoggingMixin, generics.RetrieveUpdateDestroyAPIView):
    """ FeatureMapping Feature single object management """
    queryset = Feature.objects.all()
    serializer_class = FeatureSerializer


class FeatureMappingFeatureTableView(LoggingMixin, DefaultNameOrdering, generics.ListAPIView):
    """ FeatureMapping Feature table view formatted for DataTable """
    queryset = Feature.objects.all()
    serializer_class = FeatureSerializer
    filterset_fields = ['name']

    def get(self, request, *args, **kwargs):
        return get_datatable_json(self)


# Scenario views
class FeatureMappingScenarioView(LoggingMixin, DefaultNameOrdering, generics.ListCreateAPIView):
    """
        get: List FeatureMapping Test Scenario objects according to filters
        post: Create FeatureMapping Test Scenario object
    """
    queryset = TestScenario.objects.all()
    serializer_class = TestScenarioSerializer
    filterset_fields = ['name']


class FeatureMappingScenarioDetailsView(LoggingMixin, generics.RetrieveUpdateDestroyAPIView):
    """ FeatureMapping Test Scenario single object management """
    queryset = TestScenario.objects.all()
    serializer_class = TestScenarioSerializer


class FeatureMappingScenarioTableView(LoggingMixin, DefaultNameOrdering, generics.ListAPIView):
    """ FeatureMapping Test Scenario table view formatted for DataTable """
    queryset = TestScenario.objects.all()
    serializer_class = TestScenarioSerializer
    filterset_fields = ['name']

    def get(self, request, *args, **kwargs):
        return get_datatable_json(self)


# for development needs
def feature_mapping_form(request):
    form = FeatureMappingFileForm()
    return render(request, 'api/feature_mapping_import.html', {'form': form})


def check_total(
        total: Optional[Union[int,str]],
        test_ids: Optional[Union[int,str]],
        ERROR_IN_ROW: str
    ) -> Tuple[Optional[int], List[Dict[str,str]]]:
    errors = []

    if not total:
        if not test_ids:
            errors.append({ERROR_IN_ROW: 'Missing total value for scenario without test ids'})
    else:
        try:
            total = int(total)
            if total <= 0:
                errors.append({ERROR_IN_ROW: 'Total value should be positive'})

            if test_ids:
                if total != len(test_ids.split(',')):
                    errors.append({ERROR_IN_ROW: 'Total value does not match number of test ids'})
        except ValueError:
            errors.append({ERROR_IN_ROW: 'Non-integer total value'})
            total = None

    return total, errors


def import_feature_mapping(file, serializer):
    WORKBOOK_FORMAT_ERROR = 'There must be 5 columns with "milestone", "feature", "scenario", "ids", and "total" data (column headers do not matter)'
    errors = []
    data = serializer.data

    try:
        wb = load_workbook(file)
    except Exception as e:
        message = getattr(e, 'message', repr(e))
        log.warning(f'Failed to open workbook: {message}')
        errors.append({'workbook error': message})
        return errors

    sheet = wb[wb.sheetnames[0]]
    if sheet is None:
        errors.append({'workbook error': 'No sheets found'})
        return errors

    rows = sheet.rows
    if not rows:
        errors.append({'workbook format error': WORKBOOK_FORMAT_ERROR})
        errors.append({'workbook error': 'No rows found'})
        return errors

    first_row = next(rows)
    headings = [c.value for c in first_row]
    if len(headings) != 5:
        errors.append({'workbook format error': WORKBOOK_FORMAT_ERROR})
        return errors

    fm_rules = []
    try:
        with transaction.atomic():
            mapping = FeatureMapping.objects.create(name=data['name'], owner_id=data['owner'], codec_id=data['codec'],
                                                    component_id=data['component'], platform_id=data['platform'],
                                                    os_id=data['os'])

            for i, row in enumerate(rows):
                ERROR_IN_ROW = f'workbook error (row {i + 2})'
                milestone_name, feature_name, scenario_name, ids_value, total = [cell.value for cell in row]
                total, total_errors = check_total(total, ids_value, ERROR_IN_ROW)
                errors.extend(total_errors)

                if all([milestone_name, feature_name, scenario_name]):
                    milestone, _ = Milestone.objects.get_or_create(name=milestone_name)
                    feature, _ = Feature.objects.get_or_create(name=feature_name)
                    scenario, _ = TestScenario.objects.get_or_create(name=scenario_name)

                    fm_rules.append(FeatureMappingRule(
                        mapping=mapping, milestone=milestone, feature=feature, scenario=scenario,
                        ids=ids_value if ids_value else None, total=total
                    ))
                else:
                    errors.append({ERROR_IN_ROW: 'Empty cells'})
            FeatureMappingRule.objects.bulk_create(fm_rules)

            # check rules for conflicts
            issues = defaultdict(list)
            scenario_names = FeatureMappingRule.objects.filter(mapping=mapping) \
                .values_list('scenario__name', flat=True).distinct()

            for name in scenario_names:
                rules = FeatureMappingRule.objects.filter(mapping_id=mapping, scenario__name=name)
                ids_list = list(rules.values_list('ids', flat=True))

                issues = check_rules_conflicts(ids_list, name, issues)

            if issues:
                raise ValidationException(dict(issues))
    except IntegrityError as e:
        if 'duplicate key' in str(e):
            errors.append({'Import error': 'Rule duplicate creation attempt'})
    except ValidationException as e:
        errors.append(e.args[0])

    return errors
