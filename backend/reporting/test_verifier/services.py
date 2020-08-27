from dataclasses import dataclass
from typing import List

from openpyxl import load_workbook

from api.models import Platform

from test_verifier.models import Codec, Feature, SubFeature, FeatureCategory

from api.models import Generation

PLATFORM_FEATURE_SUPPORT = 'platform feature support'

# index of the row with main headers, words Codec, Feature, Feature Category and etc. located there
LOWEST_HEADER_ROW = 3

COL_NAME_MAPPING = {
    'codec': 'Codec',
    'feature category': 'FeatureCategory',
    'feature': 'Feature',
    'sub feature': 'SubFeature',
    'notes': 'Notes',
}

YES = 'Y'
NO = 'N'


def import_features(file, user):
    outcome = OutcomeBuilder()
    platform_cells, column_mapping, sheet = verify_file(file, outcome)

    if not outcome.is_success():
        return outcome

    platform_builder = PlatformBuilder(sheet, platform_cells['row'] + 1,
                                       platform_cells['start'], platform_cells['end'], outcome)
    platform_builder.build()

    if not outcome.is_success():
        return outcome

    subfeature_builder = SubFeatureBuilder(sheet, LOWEST_HEADER_ROW, column_mapping,
                                           outcome, platform_builder.platforms, user)
    subfeature_builder.parse_subfeatures()
    return outcome


def verify_file(file, outcome):
    try:
        workbook = load_workbook(file)
    except Exception as e:
        message = getattr(e, 'message', repr(e))
        outcome.add_workbook_error(message)
        return None, None, None

    # TODO implement more complex sheet selection
    sheet = workbook.worksheets[0]

    rows = tuple(sheet.rows)[:3]
    column_mapping = dict()
    platform_cells = {}
    # cell_ranges = [cols for cols in sheet.merged_cells.ranges]
    cell_ranges = sheet.merged_cell_ranges
    for row in rows:
        for cell in row:
            if cell.value is None:
                continue
            if cell.value.lower() == PLATFORM_FEATURE_SUPPORT:
                for cell_range in cell_ranges:  # find out cell range which starts with current cell
                    if cell_range.min_col == cell.column and cell_range.min_row == cell.row:
                        platform_cells = {'row': cell.row,
                                          'start': cell_range.min_col,
                                          'end': cell_range.max_col}
                        break
            else:
                mapped_name = COL_NAME_MAPPING.get(str(cell.value).lower(), None)
                if mapped_name is not None:
                    column_mapping[mapped_name] = cell.column - 1
    return platform_cells, column_mapping, sheet


@dataclass
class PlatformRecord:
    gen: str = ''
    win_col: int = None
    lin_col: int = None

    @property
    def api_gen_name(self):
        """ To map M11 from the excel sheet to Gen11 from the DB """
        return self.gen.replace('M', 'Gen')


class SubFeatureBuilder:
    def __init__(self, sheet, row_number, column_mapping, outcome, platforms: List[PlatformRecord], user):
        self.sheet = sheet
        self.row_index = row_number
        self.column_mapping = column_mapping
        self.outcome = outcome
        self.platforms = platforms
        self.records = []
        self.entities = []
        self.user = user

    def parse_subfeatures(self):
        rows = tuple(self.sheet.rows)
        for row in rows[self.row_index:]:
            record = {'Codec': None,
                      'FeatureCategory': None,
                      'Feature': None,
                      'SubFeature': None,
                      'Notes': None,
                      'lin_platforms': [],
                      'win_platforms': [],
                      'id': None}
            for col_name, col_number in self.column_mapping.items():
                record[col_name] = row[col_number].value
                if record[col_name]:
                    record[col_name] = str(record[col_name]).strip()
            for platform in self.platforms:
                if platform.lin_col and row[platform.lin_col].value == YES:
                    record['lin_platforms'].append(platform)
                if platform.win_col and row[platform.win_col].value == YES:
                    record['win_platforms'].append(platform)
            self._create_dependencies(record)  # create Codec, Feature, FeatureCategory objects if necessary
            self._create_if_needed(record)  # create subfeature object if needed
            self.records.append(record)

    def _create_dependencies(self, record):
        for model in (Codec, FeatureCategory, Feature):
            model.objects.get_or_create(name=record[model.__name__])

    def _create_if_needed(self, record):
        entity = self._generate_object(record)
        try:
            record['id'] = SubFeature.objects.get(name=entity['name'],
                                                  codec=entity['codec'],
                                                  category=entity['category'],
                                                  feature=entity['feature'],
                                                  notes=entity['notes'])
        except (SubFeature.DoesNotExist, SubFeature.MultipleObjectsReturned):
            record = SubFeature(name=entity['name'],
                                codec=entity['codec'],
                                category=entity['category'],
                                feature=entity['feature'],
                                notes=entity['notes'],
                                imported=True,
                                created_by=self.user)
            record.save()
            for platform in entity['lin_platforms']:
                record.lin_platforms.add(platform)
            for platform in entity['win_platforms']:
                record.win_platforms.add(platform)
            record.save()

    def _generate_object(self, record):
        """ Generate dict which contains all fields needed to create new subfeature"""

        codec = Codec.objects.get(name=record['Codec'])
        feature_category = FeatureCategory.objects.get(name=record['FeatureCategory'])
        feature = Feature.objects.get(name=record['Feature'])
        lin_platforms = list(Platform.objects
                             .filter(generation__name__in=[pl.api_gen_name for pl in
                                                           record['lin_platforms']]).values_list('id', flat=True))
        win_platforms = list(Platform.objects
                             .filter(generation__name__in=[pl.api_gen_name for pl in
                                                           record['win_platforms']]).values_list('id', flat=True))
        return {'name': record['SubFeature'],
                'codec': codec,
                'category': feature_category,
                'feature': feature,
                'lin_platforms': lin_platforms,
                'win_platforms': win_platforms,
                'notes': record['Notes']}


class OutcomeBuilder:
    def __init__(self):
        self.success = False
        self.errors = list()
        self.warnings = dict()
        self.changes = dict(added=0, updated=0, skipped=0)

    def build(self):
        outcome = {}
        success = self.is_success()
        if success:
            outcome = dict(
                success=success,
                changes=self.changes,
            )
        else:
            outcome = dict(
                success=success,
                errors=self.errors,
                warnings=self.warnings,
            )

        return outcome

    def is_success(self):
        return len(self.errors) + len(self.warnings) == 0

    def add_warning(self, warning):
        self.warnings[warning] = self.warnings.get(warning, 0) + 1

    def __add_error(self, code, message, **kwargs):
        err = dict(code=code, message=message, **kwargs)
        if err not in self.errors:
            self.errors.append(err)

    def add_missing_field_error(self, missing_field):
        model_name, fields = missing_field

        err = dict(
            code='ERR_MISSING_ENTITY',
            message=f'Missing field {missing_field}',
            entity=dict(model=model_name, fields=fields),
        )

        self.add_warning(f'{model_name} with properties {fields} does not exist.')

        if err not in self.errors:
            self.errors.append(err)

    def add_workbook_error(self, message):
        self.__add_error('ERR_WORKBOOK_EXCEPTION', message)


class PlatformBuilder:
    def __init__(self, sheet, start_row, start_col, end_col, outcome):
        self.sheet = sheet
        self.start_row = start_row
        self.start_col = start_col
        self.end_col = end_col
        self.outcome = outcome
        self.platforms: List[PlatformRecord] = []

    def build(self):
        rows = tuple(self.sheet.rows)

        # row where should be list of platforms, some cells are merged
        # because there are os mapping on the row below
        platform_lever_row = rows[self.start_row - 1]

        # row contains os families mapped on every platform
        os_lever_row = rows[self.start_row]

        # set of columns numbers where platform are located
        columns_to_visit = set(col_num for col_num in range(self.start_col - 1, self.end_col))
        columns_ranges = self.sheet.merged_cells.ranges  # ranges of merged cells
        for col_range in columns_ranges:  # iterate through merged cells to find out platforms
            # if cell range is located in platform space
            if (col_range.min_row == self.start_row == col_range.max_row
                    and col_range.min_col >= self.start_col and col_range.max_col <= self.end_col):
                platform = PlatformRecord()

                # obtain platform gen name - value of left cell in cell range, example - "M11"
                platform.gen = platform_lever_row[col_range.min_col - 1].value.strip()

                # get numbers of columns which contains OS mapping (win/lin for platforms)
                for attr_name, os_group in zip(('lin_col', 'win_col'), ('linux', 'windows')):
                    if os_lever_row[col_range.min_col - 1].value.strip().lower() == os_group:
                        setattr(platform, attr_name, col_range.min_col - 1)
                    elif os_lever_row[col_range.max_col - 1].value.strip().lower() == os_group:
                        setattr(platform, attr_name, col_range.max_col - 1)

                columns_to_visit.discard(col_range.min_col - 1)
                columns_to_visit.discard(col_range.max_col - 1)
                self.platforms.append(platform)

        # handle platforms which support only linux or windows
        for column_number in columns_to_visit:
            platform = PlatformRecord()
            platform.gen = platform_lever_row[column_number].value

            for attr_name, os_group in zip(('lin_col', 'win_col'), ('linux', 'windows')):
                if os_lever_row[column_number].value.lower() == os_group:
                    setattr(platform, attr_name, column_number)
            self.platforms.append(platform)

        self._handle_new_gens()

    def _handle_new_gens(self):
        for platform in self.platforms:
            try:
                found_gen = Generation.objects.get(name=platform.api_gen_name)
            except (Generation.DoesNotExist, Generation.MultipleObjectsReturned):
                self.outcome.add_missing_field_error((Generation.__name__, {'name': platform.api_gen_name}))
